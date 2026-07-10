"""
Render publication tables (CSV + formatted .xlsx) and figures (300 dpi PNG + SVG)
from the outputs/tables/ CSVs produced by scripts/05 and scripts/06.

Table/figure presentation follows the two reference papers in refs/: plain bordered
tables with a light-gray header band and N stated in the table title (matching the
GLP-1RA FAERS paper's Table 3 style), and simple, uncluttered bar/scatter charts
with data labels rather than dense chart chrome (matching both references' Figure 2
/ Fig 2 style). Colors follow the categorical/sequential/diverging rules in the
project's dataviz skill (see conversation) -- one hue per series, diverging
blue/red only for the ROR-centered-at-1 heatmap, status-style highlight (not a
generic categorical color) for "signal vs not" in the volcano plot.

Every table title and figure caption states N and the data window
(2025Q1-2026Q1, first approval 2025-01-17) per project standing rule.
"""

from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
PROCESSED_DIR = REPO_ROOT / "data" / "processed"
TABLES_DIR = REPO_ROOT / "outputs" / "tables"
FIGURES_DIR = REPO_ROOT / "outputs" / "figures"

N_CASES = 416
DATA_WINDOW = "2025Q1-2026Q1"
WINDOW_NOTE = f"N={N_CASES} cases, datopotamab deruxtecan Primary Suspect, FAERS {DATA_WINDOW}"

# --- dataviz palette (see project dataviz skill; validated categorical/sequential/diverging) ---
COLOR_SERIES1_BLUE = "#2a78d6"
COLOR_SERIES6_RED = "#e34948"
COLOR_SERIES8_ORANGE = "#eb6834"
COLOR_MUTED = "#898781"
COLOR_GRIDLINE = "#e1e0d9"
COLOR_TEXT_PRIMARY = "#0b0b0b"
COLOR_TEXT_SECONDARY = "#52514e"

plt.rcParams.update({
    "font.family": "sans-serif",
    "axes.edgecolor": COLOR_GRIDLINE,
    "axes.labelcolor": COLOR_TEXT_SECONDARY,
    "xtick.color": COLOR_TEXT_SECONDARY,
    "ytick.color": COLOR_TEXT_SECONDARY,
    "text.color": COLOR_TEXT_PRIMARY,
    "axes.grid": True,
    "grid.color": COLOR_GRIDLINE,
    "grid.linewidth": 0.6,
    "axes.axisbelow": True,
})

# Label-anchored PT sets (see scripts/04, scripts/06 for full curation rationale).
AESI_LABEL_PTS = {
    "Interstitial lung disease", "Pneumonitis",
    "Dry eye", "Keratitis", "Ulcerative keratitis", "Conjunctivitis",
    "Corneal epitheliopathy", "Corneal erosion", "Eye irritation", "Eye pain",
    "Eye pruritus", "Ocular discomfort", "Ocular hyperaemia", "Ocular toxicity",
    "Vision blurred", "Abnormal sensation in eye",
    "Stomatitis", "Mucosal inflammation",
    "Infusion related reaction", "Infusion related hypersensitivity reaction",
    "Infusion site extravasation", "Infusion site pain", "Infusion site rash",
}
# Label's "most common adverse reactions >=20%, NSCLC pool" (docs/candidate_justification.md),
# restricted to PTs that actually string-match our data (verified by hand, not guessed).
COMMON_AR_LABEL_PTS = {
    "Nausea", "Alopecia", "Fatigue", "Constipation", "Decreased appetite",
    "Musculoskeletal pain", "Rash", "White blood cell count decreased",
}
LABEL_LISTED_PTS = AESI_LABEL_PTS | COMMON_AR_LABEL_PTS


# ============================================================================
# xlsx formatting helpers
# ============================================================================

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def write_sheet(wb: Workbook, sheet_name: str, title: str, df: pd.DataFrame, section_col: str = None):
    display_cols = [c for c in df.columns if c != section_col]
    ws = wb.create_sheet(sheet_name[:31])
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(display_cols), 1))

    header_row = 3
    for j, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=str(col))
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for i, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        is_section = section_col is not None and bool(row.get(section_col, False))
        for j, col in enumerate(display_cols, start=1):
            val = row[col]
            if isinstance(val, (np.floating, float)) and pd.notna(val):
                val = round(float(val), 3)
            cell = ws.cell(row=i, column=j, value=(None if pd.isna(val) else val))
            cell.border = THIN_BORDER
            if is_section:
                cell.font = Font(bold=True)
                cell.fill = SECTION_FILL

    for j, col in enumerate(display_cols, start=1):
        width = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 10) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(width, 45)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws


def save_table(name: str, title: str, df: pd.DataFrame, wb: Workbook, section_col: str = None):
    df.to_csv(TABLES_DIR / f"{name}.csv", index=False)
    write_sheet(wb, name, title, df, section_col=section_col)
    print(f"  {name}: {len(df)} rows -> outputs/tables/{name}.csv (+ xlsx sheet)")


# ============================================================================
# T1: demographic & reporting characteristics
# ============================================================================

def build_table1() -> pd.DataFrame:
    sections = []

    def add_section(label, df, item_col, n_col, pct_col=None):
        sections.append({"item": label, "n": None, "pct": None, "is_section": True})
        for _, r in df.iterrows():
            pct = r[pct_col] if pct_col else round(100 * r[n_col] / N_CASES, 1)
            sections.append({"item": f"  {r[item_col]}", "n": int(r[n_col]), "pct": pct, "is_section": False})

    sex = pd.read_csv(TABLES_DIR / "demo_sex.csv")
    add_section("Sex", sex, "sex", "n", "pct")

    age = pd.read_csv(TABLES_DIR / "demo_age_band.csv")
    add_section("Age band", age, "age_band", "n", "pct")

    reporter = pd.read_csv(TABLES_DIR / "demo_reporter_type.csv")
    add_section("Reporter type", reporter, "reporter_type", "n", "pct")

    year = pd.read_csv(TABLES_DIR / "demo_report_year.csv")
    add_section("Report year (init. FDA receipt)", year, "report_year", "n")

    country = pd.read_csv(TABLES_DIR / "demo_country.csv")
    reporter_country = country[country["country_type"] == "reporter"].sort_values("n", ascending=False).head(8)
    add_section("Reporting country (top 8, reporter_country)", reporter_country, "country", "n", "pct")

    serious = pd.read_csv(TABLES_DIR / "serious_outcomes.csv")
    add_section("Serious outcome (FAERS OUTC code)", serious, "label", "n_cases", "pct_of_all_cases")

    df = pd.DataFrame(sections)[["item", "n", "pct"]]
    return df


# ============================================================================
# T2: top ~15 PTs by report frequency
# ============================================================================

def build_table2(signals_all: pd.DataFrame) -> pd.DataFrame:
    top = signals_all.sort_values("a", ascending=False).head(15).copy()
    top["label_listed"] = top["pt"].isin(LABEL_LISTED_PTS).map({True: "Y", False: "N"})
    out = top[["pt", "signal_category", "a", "ror", "ror_ci_lower", "ror_ci_upper", "prr", "ic", "ic025", "ebgm", "eb05", "label_listed"]].copy()
    out.columns = ["PT", "Signal category", "a (n cases)", "ROR", "ROR 95% CI lower", "ROR 95% CI upper",
                   "PRR", "IC", "IC025", "EBGM", "EB05", "Label-listed (Y/N)"]
    return out.round(3)


# ============================================================================
# T3: top signals by strength across all four algorithms
# ============================================================================

def build_table3(signals_significant: pd.DataFrame) -> pd.DataFrame:
    # Grouped, not a flat ROR-sorted list: Clinical AE signals first, then
    # Administrative/case-context codes, each internally sorted by ROR. Mixing
    # the two in one ranked list was the exact "unclean" complaint this
    # grouping addresses -- a reader scanning top-to-bottom by ROR alone would
    # see "Disease progression" (an artifact) ranked above "Interstitial lung
    # disease" (a real AE) with no visual cue they're different kinds of thing.
    category_order = pd.Categorical(
        signals_significant["signal_category"], categories=["Clinical AE", "Administrative/case-context"]
    )
    out = signals_significant.assign(_cat=category_order).sort_values(
        ["_cat", "ror"], ascending=[True, False]
    ).copy()
    out = out[["pt", "signal_category", "a", "ror", "ror_ci_lower", "ror_ci_upper", "prr", "chi2",
               "ic", "ic025", "ebgm", "eb05", "strict_signal"]].copy()
    out.columns = ["PT", "Signal category", "a (n cases)", "ROR", "ROR 95% CI lower", "ROR 95% CI upper",
                   "PRR", "chi2 (Yates)", "IC", "IC025", "EBGM", "EB05", "Strict signal (+EB05>2)"]
    return out.round(3)


# ============================================================================
# T4: signal distribution by SOC
# ============================================================================

def build_table4(soc_rollup: pd.DataFrame) -> pd.DataFrame:
    out = soc_rollup.sort_values("n_consensus_signals", ascending=False).copy()
    out.columns = ["System Organ Class", "n PTs", "n consensus signals", "Total cases (sum of a)"]
    return out


# ============================================================================
# T5: time-to-onset summary
# ============================================================================

def build_table5(tto_summary: pd.DataFrame) -> pd.DataFrame:
    out = tto_summary.copy()
    out["group"] = out.apply(
        lambda r: f"{r['group']} *" if r.get("low_n_caveat") else r["group"], axis=1
    )
    out = out[["group", "n", "median_days", "q1_days", "q3_days",
               "weibull_shape", "shape_ci_lower", "shape_ci_upper", "hazard_pattern"]]
    out.columns = ["Group", "n", "Median (days)", "Q1 (days)", "Q3 (days)",
                   "Weibull shape (beta)", "beta 95% CI lower", "beta 95% CI upper", "Hazard pattern"]
    return out.round(3)


# ============================================================================
# Figures
# ============================================================================

def savefig(fig, name, caption, bottom=0.22):
    # Reserve a dedicated caption band below the axes so it never overlaps the
    # x-axis label/ticks -- bbox_inches="tight" alone isn't enough since it just
    # tightens around existing artists, it doesn't stop them from overlapping
    # each other first. `bottom` is overridable for figures with long rotated
    # tick labels (e.g. F5) that need more room than the default.
    fig.subplots_adjust(bottom=bottom)
    fig.text(0.01, 0.02, caption, fontsize=6.5, color=COLOR_TEXT_SECONDARY, wrap=True, va="bottom")
    fig.savefig(FIGURES_DIR / f"{name}.png", dpi=300, bbox_inches="tight", pad_inches=0.3)
    fig.savefig(FIGURES_DIR / f"{name}.svg", bbox_inches="tight", pad_inches=0.3)
    plt.close(fig)
    print(f"  {name}: outputs/figures/{name}.png (300dpi) + .svg")


def figure1_soc_bar(soc_rollup: pd.DataFrame):
    df = soc_rollup.sort_values("n_consensus_signals", ascending=False)
    short_labels = [s.split(" (")[0] for s in df["soc"]]  # drop long parenthetical for the axis
    fig, ax = plt.subplots(figsize=(7, 4.5))
    bars = ax.bar(range(len(df)), df["n_consensus_signals"], color=COLOR_SERIES1_BLUE, width=0.5)
    ax.set_xticks(range(len(df)))
    rotate = 15 if len(df) > 3 else 0
    ax.set_xticklabels(short_labels, rotation=rotate, ha="right" if rotate else "center", fontsize=9)
    ax.set_ylabel("Consensus signals (n)")
    ax.set_ylim(0, df["n_consensus_signals"].max() * 1.2)
    ax.set_title("F1. Consensus signals by MedDRA System Organ Class", fontsize=11, loc="left")
    for rect, val in zip(bars, df["n_consensus_signals"]):
        ax.text(rect.get_x() + rect.get_width() / 2, rect.get_height() + 0.15, str(int(val)),
                 ha="center", va="bottom", fontsize=9, color=COLOR_TEXT_PRIMARY)
    ax.spines[["top", "right"]].set_visible(False)
    if len(df) == 1:
        ax.text(
            0.5, 0.5,
            "No licensed MedDRA PT->SOC crosswalk is available (MSSO-licensed;\n"
            "see docs/environment.md). All PTs fall into a single Unmapped bucket.",
            transform=ax.transAxes, ha="center", va="center", fontsize=9, color="white",
        )
    caption = f"F1. {WINDOW_NOTE}. SOC rollup of the 8 consensus-signal PTs from scripts/05."
    savefig(fig, "F1_soc_signal_counts", caption)


def figure2_volcano(signals_all: pd.DataFrame):
    df = signals_all.copy()
    df["log_ror"] = np.log2(df["ror"].clip(lower=1e-6))
    fig, ax = plt.subplots(figsize=(7.5, 5.5))

    non_signal = df[~df["consensus_signal"]]
    clinical = df[df["consensus_signal"] & (df["signal_category"] == "Clinical AE")]
    admin = df[df["consensus_signal"] & (df["signal_category"] == "Administrative/case-context")]

    ax.scatter(non_signal["a"], non_signal["log_ror"], s=18, color=COLOR_MUTED, alpha=0.55,
               linewidths=0, label="Not a consensus signal")
    ax.scatter(admin["a"], admin["log_ror"], s=36, color=COLOR_SERIES8_ORANGE, alpha=0.9,
               linewidths=0, label="Consensus signal -- administrative/case-context (not an AE)")
    ax.scatter(clinical["a"], clinical["log_ror"], s=36, color=COLOR_SERIES6_RED, alpha=0.9,
               linewidths=0, label="Consensus signal -- clinical AE")
    ax.axhline(0, color=COLOR_MUTED, linewidth=1)
    ax.axvline(3, color=COLOR_MUTED, linewidth=1)
    ax.set_xscale("log")
    ax.set_xlabel("Report count (a, log scale)")
    ax.set_ylabel("log2(ROR)")
    ax.set_title("F2. Signal volcano plot: ROR vs. report count", fontsize=11, loc="left")
    top_labels = pd.concat([clinical, admin]).sort_values("a").reset_index(drop=True)
    for i, r in top_labels.iterrows():
        # Alternate the label above/below its point so close-together points
        # (e.g. Dry eye / Infusion related reaction, similar a and ROR) don't
        # overlap -- a fixed offset for every label collided there.
        dy = 8 if i % 2 == 0 else -12
        va = "bottom" if dy > 0 else "top"
        ax.annotate(r["pt"], (r["a"], r["log_ror"]), fontsize=7, color=COLOR_TEXT_PRIMARY,
                    xytext=(4, dy), textcoords="offset points", va=va)
    ax.legend(frameon=False, fontsize=8, loc="upper left")
    ax.spines[["top", "right"]].set_visible(False)
    caption = (f"F2. {WINDOW_NOTE}. Reference lines: log2(ROR)=0 (null value) and "
               "a=3 (Evans minimum case count). Consensus signals split by category -- "
               "orange points meet the statistical criteria but are FAERS administrative/"
               "case-context codes, not adverse events (see Discussion).")
    savefig(fig, "F2_volcano", caption)


def figure3_forest(signals_significant: pd.DataFrame):
    # Two grouped blocks (Clinical AE, then Administrative/case-context), each
    # sorted by ROR, with a visual gap and distinct marker color between them --
    # a single ROR-sorted list mixed real AEs with FAERS coding artifacts with
    # no visual distinction, which was the source of the "unclean" complaint.
    clinical = signals_significant[signals_significant["signal_category"] == "Clinical AE"].sort_values("ror")
    admin = signals_significant[signals_significant["signal_category"] == "Administrative/case-context"].sort_values("ror")

    gap = 1  # blank row between groups
    n_rows = len(clinical) + len(admin) + gap
    fig, ax = plt.subplots(figsize=(7.5, 0.5 * n_rows + 1.5))

    y_admin = np.arange(len(admin))
    y_clinical = np.arange(len(admin) + gap, len(admin) + gap + len(clinical))

    for sub_df, y_pos, color, label in [
        (admin, y_admin, COLOR_SERIES8_ORANGE, "Administrative/case-context"),
        (clinical, y_clinical, COLOR_SERIES6_RED, "Clinical AE"),
    ]:
        if len(sub_df) == 0:
            continue
        ax.errorbar(sub_df["ror"], y_pos,
                    xerr=[sub_df["ror"] - sub_df["ror_ci_lower"], sub_df["ror_ci_upper"] - sub_df["ror"]],
                    fmt="o", color=color, ecolor=color, elinewidth=1.5, capsize=3, markersize=6, label=label)

    all_y = list(y_admin) + list(y_clinical)
    all_pt = list(admin["pt"]) + list(clinical["pt"])
    ax.axvline(1, color=COLOR_MUTED, linewidth=1)
    ax.axhline(len(admin) + gap / 2 - 0.5, color=COLOR_GRIDLINE, linewidth=1)
    ax.set_yticks(all_y)
    ax.set_yticklabels(all_pt, fontsize=9)
    ax.set_xscale("log")
    ax.set_xlabel("ROR (log scale, 95% CI)")
    ax.set_title("F3. Forest plot of consensus signals, grouped by category", fontsize=11, loc="left")
    ax.legend(frameon=False, fontsize=8, loc="lower right")
    ax.spines[["top", "right"]].set_visible(False)
    caption = (f"F3. {WINDOW_NOTE}. All 8 PTs meeting the consensus signal rule (ROR CI-lower>1 AND "
               "Evans PRR/chi2 AND IC025>0), grouped by whether the PT is a clinical adverse event "
               "or a FAERS administrative/case-context code (see Discussion) -- not one ROR-ranked list.")
    savefig(fig, "F3_forest_consensus", caption)


def figure4_tto_histogram():
    tto = pd.read_csv(TABLES_DIR / "tto_case_level.csv")
    summary = pd.read_csv(TABLES_DIR / "tto_summary.csv")
    overall = summary[summary["group"] == "Overall"].iloc[0]

    fig, ax = plt.subplots(figsize=(7, 4.5))
    days = tto["onset_days"].clip(lower=0.5)
    ax.hist(days, bins=20, color=COLOR_SERIES1_BLUE, alpha=0.75, density=True, label=f"Observed onset (n={len(days)})")

    if pd.notna(overall["weibull_shape"]):
        from scipy.stats import weibull_min
        shape = overall["weibull_shape"]
        median = overall["median_days"]
        # lifelines WeibullFitter parameterizes scale (lambda_) s.t. median = lambda_*(ln2)^(1/rho);
        # recover an approximate scale from the reported median/shape for plotting the fitted curve.
        scale = median / (np.log(2) ** (1 / shape))
        x = np.linspace(0.5, days.max(), 200)
        ax.plot(x, weibull_min.pdf(x, c=shape, scale=scale), color=COLOR_SERIES6_RED, linewidth=2,
                label=f"Fitted Weibull (beta={shape:.2f})")

    ax.set_xlabel("Time to onset (days)")
    ax.set_ylabel("Density")
    ax.set_title("F4. Time-to-onset distribution with fitted Weibull curve", fontsize=11, loc="left")
    ax.legend(frameon=False, fontsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    caption = (f"F4. {WINDOW_NOTE}. Onset = EVENT_DT - drug-specific START_DT; valid for "
               f"{len(days)}/{N_CASES} cases (21.6%) after dropping negative/implausible values (see docs).")
    savefig(fig, "F4_tto_histogram_weibull", caption)


def figure5_subgroup_heatmap(signals_significant: pd.DataFrame):
    # Restricted to Clinical AE signals only -- subgroup breakdown of FAERS
    # administrative/case-context codes (Disease progression, Off label use,
    # etc.) by sex/age isn't a clinically meaningful question, and including
    # them alongside real AEs was part of why this figure read as cluttered.
    clinical_pts = set(
        signals_significant.loc[signals_significant["signal_category"] == "Clinical AE", "pt"]
    )
    sex_df = pd.read_csv(TABLES_DIR / "subgroup_ror_sex.csv")
    age_df = pd.read_csv(TABLES_DIR / "subgroup_ror_age_band.csv")
    sex_df = sex_df[sex_df["pt"].isin(clinical_pts)].rename(columns={"sex_group": "group"})
    age_df = age_df[age_df["pt"].isin(clinical_pts)].rename(columns={"age_band": "group"})
    combined = pd.concat([sex_df.assign(dim="Sex"), age_df.assign(dim="Age band")], ignore_index=True)

    pts = combined["pt"].unique()
    groups = combined[["dim", "group"]].drop_duplicates()
    groups["label"] = groups["dim"] + ": " + groups["group"]
    pivot = combined.pivot_table(index="pt", columns=["dim", "group"], values="ror")
    pivot.columns = [f"{d}: {g}" for d, g in pivot.columns]
    pivot = pivot.reindex(pts)

    log_pivot = np.log2(pivot.clip(lower=1e-6))
    vmax = np.nanmax(np.abs(log_pivot.values)) if np.isfinite(log_pivot.values).any() else 1

    fig, ax = plt.subplots(figsize=(0.9 * len(pivot.columns) + 3, 0.4 * len(pivot) + 2))
    im = ax.imshow(log_pivot.values, cmap="RdBu_r", vmin=-vmax, vmax=vmax, aspect="auto")
    ax.set_xticks(range(len(pivot.columns)))
    ax.set_xticklabels(pivot.columns, rotation=45, ha="right", fontsize=8)
    ax.set_yticks(range(len(pivot)))
    ax.set_yticklabels(pivot.index, fontsize=8)
    for i in range(len(pivot)):
        for j in range(len(pivot.columns)):
            val = pivot.values[i, j]
            if pd.notna(val):
                ax.text(j, i, f"{val:.1f}", ha="center", va="center", fontsize=6.5,
                        color="white" if abs(np.log2(max(val, 1e-6))) > vmax * 0.5 else COLOR_TEXT_PRIMARY)
    cbar = fig.colorbar(im, ax=ax, shrink=0.7)
    cbar.set_label("log2(ROR)", fontsize=8)
    ax.set_title("F5. Subgroup ROR heatmap (clinical AE signals x sex / age band)", fontsize=11, loc="left")
    caption = (f"F5. {WINDOW_NOTE}. Restricted to the 4 Clinical AE consensus signals -- "
               "Administrative/case-context codes (Disease progression, Off label use, etc.) "
               "are excluded, since a sex/age breakdown of coding artifacts isn't clinically "
               "meaningful. Diverging scale centered at ROR=1 (white); blue=ROR<1, red=ROR>1. "
               "Cells with Haldane-Anscombe continuity correction and 'Unknown'-group concentration "
               "are not distinguished here -- see subgroup_ror_*.csv for that caveat.")
    # bottom margin is a fraction of total figure height, and restricting to 4
    # clinical PTs (down from 8) shrank the figure -- the same fraction now
    # gives less absolute room for the rotated tick labels + caption, so it's
    # bumped up here rather than left at the 8-row value.
    savefig(fig, "F5_subgroup_heatmap", caption, bottom=0.48)


# ============================================================================
# Main
# ============================================================================

def main():
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)

    signals_all = pd.read_csv(TABLES_DIR / "signals_all.csv")
    signals_significant = pd.read_csv(TABLES_DIR / "signals_significant.csv")
    soc_rollup = pd.read_csv(TABLES_DIR / "soc_rollup.csv")
    tto_summary = pd.read_csv(TABLES_DIR / "tto_summary.csv")

    print("Building tables...")
    wb = Workbook()
    wb.remove(wb.active)

    t1 = build_table1()
    t1_title = f"Table 1. Demographic & reporting characteristics of PS cases ({WINDOW_NOTE})"
    t1.to_csv(TABLES_DIR / "T1_demographics.csv", index=False)
    t1_for_xlsx = t1.copy()
    t1_for_xlsx["is_section"] = t1_for_xlsx["n"].isna()
    write_sheet(wb, "T1_demographics", t1_title, t1_for_xlsx, section_col="is_section")
    print(f"  T1_demographics: {len(t1)} rows -> outputs/tables/T1_demographics.csv (+ xlsx sheet)")

    t2 = build_table2(signals_all)
    save_table("T2_top_pts_by_frequency", f"Table 2. Top 15 PTs by report frequency ({WINDOW_NOTE})", t2, wb)

    t3 = build_table3(signals_significant)
    save_table("T3_top_signals_by_strength",
               f"Table 3. Consensus signals by category (Clinical AE, then Administrative/case-context), "
               f"ranked by ROR within each ({WINDOW_NOTE})", t3, wb)

    t4 = build_table4(soc_rollup)
    save_table("T4_soc_distribution", f"Table 4. Signal distribution by MedDRA SOC ({WINDOW_NOTE})", t4, wb)

    t5 = build_table5(tto_summary)
    save_table("T5_time_to_onset", f"Table 5. Time-to-onset summary, overall and by AESI ({WINDOW_NOTE}); "
               "* = n<10, interpret with caution", t5, wb)

    wb.save(TABLES_DIR / "tables.xlsx")
    print(f"Wrote {TABLES_DIR / 'tables.xlsx'} (5 sheets)")

    print("\nBuilding figures...")
    figure1_soc_bar(soc_rollup)
    figure2_volcano(signals_all)
    figure3_forest(signals_significant)
    figure4_tto_histogram()
    figure5_subgroup_heatmap(signals_significant)

    print("\nAll tables and figures written.")


if __name__ == "__main__":
    main()
